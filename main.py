import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QLineEdit, QHBoxLayout, QGroupBox, QRadioButton
from openpyxl import Workbook, load_workbook
import datetime

class MoneyManagerApp(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.setWindowTitle('Money Manager')
        self.setGeometry(100, 100, 400, 600)

        layout = QVBoxLayout()

        # Group box for income
        income_groupbox = QGroupBox('Income')
        income_groupbox.setStyleSheet('font-size: 17px; margin: 10px; border: 1px solid black; border-radius: 5px;')
        income_layout = QVBoxLayout()
        self.income_type_label = QLabel('Type')
        self.income_type_input = QLineEdit()
        income_layout.addWidget(self.income_type_label)
        income_layout.addWidget(self.income_type_input)

        self.income_amount_label = QLabel('Amount')
        self.income_amount_input = QLineEdit()
        income_layout.addWidget(self.income_amount_label)
        income_layout.addWidget(self.income_amount_input)

        self.add_income_button = QPushButton('Add Income')
        self.add_income_button.setStyleSheet("background-color: rgb(189, 253, 108 );")
        self.add_income_button.clicked.connect(self.addIncome)
        income_layout.addWidget(self.add_income_button)

        income_groupbox.setLayout(income_layout)
        layout.addWidget(income_groupbox)

        # Group box for expense
        expense_groupbox = QGroupBox('Expense')
        expense_groupbox.setStyleSheet('font-size: 17px; margin: 10px; border: 1px solid black; border-radius: 5px;')
        expense_layout = QVBoxLayout()
        self.expense_type_label = QLabel('Type:')
        self.expense_type_input = QLineEdit()
        expense_layout.addWidget(self.expense_type_label)
        expense_layout.addWidget(self.expense_type_input)

        self.expense_amount_label = QLabel('Amount:')
        self.expense_amount_input = QLineEdit()
        expense_layout.addWidget(self.expense_amount_label)
        expense_layout.addWidget(self.expense_amount_input)

        self.add_expense_button = QPushButton('Add Expense')
        self.add_expense_button.setStyleSheet("background-color: rgb(250, 66, 66);")
        self.add_expense_button.clicked.connect(self.addExpense)
        expense_layout.addWidget(self.add_expense_button)

        expense_groupbox.setLayout(expense_layout)
        layout.addWidget(expense_groupbox)

        # Group box for saving
        saving_groupbox = QGroupBox('Saving')
        saving_groupbox.setStyleSheet('font-size: 17px; margin: 10px; border: 1px solid black; border-radius: 5px;')
        saving_layout = QVBoxLayout()
        self.saving_amount_label = QLabel('Amount:')
        self.saving_amount_input = QLineEdit()
        saving_layout.addWidget(self.saving_amount_label)
        saving_layout.addWidget(self.saving_amount_input)

        self.add_saving_button = QPushButton('Add Saving')
        self.add_saving_button.setStyleSheet("background-color: rgb(66, 214, 250);")
        self.add_saving_button.clicked.connect(self.addSaving)
        saving_layout.addWidget(self.add_saving_button)

        saving_groupbox.setLayout(saving_layout)
        layout.addWidget(saving_groupbox)

        self.setLayout(layout)

        self.show()

    def addIncome(self):
        income_type = self.income_type_input.text()
        income_amount = int(self.income_amount_input.text())

        wb = load_workbook('money2.xlsx')
        ws = wb.active
        day = datetime.datetime.now().day
        month = datetime.datetime.now().month
        months = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August',
                  9: 'September', 10: 'October', 11: 'November', 12: 'December'}
        day = str(day)
        month = months[month]

        if ws.cell(row=ws.max_row, column=6).value is not None:
            ws.cell(row=ws.max_row + 1, column=6).value = income_amount
            ws.cell(row=ws.max_row, column=5).value = f'{income_type} : {day} / {month}'
        else:
            ws.cell(row=ws.max_row, column=6).value = income_amount
            ws.cell(row=ws.max_row, column=5).value = f'{income_type} : {day} / {month}'

        ws['B8'] = ws['B8'].value + income_amount

        wb.save('money2.xlsx')

    def addExpense(self):
        expense_type = self.expense_type_input.text()
        expense_amount = int(self.expense_amount_input.text())

        wb = load_workbook('money2.xlsx')
        ws = wb.active
        day = datetime.datetime.now().day
        month = datetime.datetime.now().month
        months = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August',
                  9: 'September', 10: 'October', 11: 'November', 12: 'December'}
        day = str(day)
        month = months[month]

        if ws.cell(row=ws.max_row, column=9).value is not None:
            ws.cell(row=ws.max_row + 1, column=9).value = expense_amount
            ws.cell(row=ws.max_row , column=8).value = f'{expense_type} : {day} / {month}'
            if expense_type == 'vape':
                ws.cell(row=ws.max_row, column=12).value = expense_amount
            if expense_type == 'rappi':
                ws['M3'] = ws['M3'].value - expense_amount
        else:
            ws.cell(row=ws.max_row, column=9).value = expense_amount
            ws.cell(row=ws.max_row , column=8).value = f'{expense_type} : {day} / {month}'
            if expense_type == 'vape':
                ws.cell(row=ws.max_row, column=12).value = expense_amount
            if expense_type == 'rappi':
                ws['M3'] = ws['M3'].value - expense_amount

        ws['B8'] = int(ws['B8'].value) - expense_amount

        wb.save('money2.xlsx')

    def addSaving(self):
        saving_amount = int(self.saving_amount_input.text())

        wb = load_workbook('money2.xlsx')
        ws = wb.active

        ws['C5'] = ws['C5'].value + saving_amount
        ws['B8'] = ws['B8'].value - saving_amount

        wb.save('money2.xlsx')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MoneyManagerApp()
    sys.exit(app.exec_())
